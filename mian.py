# -*- coding: utf-8 -*-
from gensim.models import Word2Vec
from openpyxl import load_workbook, Workbook

'''
参数解读：
    LineSentence(inp)：格式简单：一句话=一行; 单词已经过预处理并被空格分隔。
    size：是每个词的向量维度； 
    window：是词向量训练时的上下文扫描窗口大小，窗口为5就是考虑前5个词和后5个词； 
    min-count：设置最低频率，默认是5，如果一个词语在文档中出现的次数小于5，那么就会丢弃； 
    workers：是训练的进程数（需要更精准的解释，请指正），默认是当前运行机器的处理器核数。这些参数先记住就可以了。
    sg ({0, 1}, optional) – 模型的训练算法: 1: skip-gram; 0: CBOW
    alpha (float, optional) – 初始学习率
    iter (int, optional) – 迭代次数，默认为5
'''

if __name__ == '__main__':
    input1 = "./Data/task2_train_reformat_cleaned.xlsx"

    output1 = "./Data/word2vec.model"
    output2 = "./Data/vector.txt"
    wb = load_workbook(input1)
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    sentences = []
    for i in range(max_row - 1):
        line = i + 2
        text = ws.cell(line, 1).value
    # with open(input1, 'r', encoding='utf8', errors='ignore') as f:
    #     for line in f:
    #     if " " in line:
        sentences.append(list(text))

    model = Word2Vec(size=300, window=5, min_count=5, workers=4)  # 定义word2vec 对象

    model.build_vocab(sentences)  # 建立初始训练集的词典
    model.train(sentences, total_examples=model.corpus_count, epochs=model.iter)  # 模型训练

    model.save(output1)  # 模型保存
    model.wv.save_word2vec_format(output2, binary=False)  # 词向量保存

'''
# 模型的训练和保存，下面两行同样可以训练，但是无法追加训练
model = Word2Vec(sentences, size=100, window=5, min_count=1, workers=4)
model.save("word2vec.model")
'''
